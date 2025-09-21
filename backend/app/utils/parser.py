import re
from typing import List, Dict, Any
from datetime import datetime
from io import BytesIO

LINE_REGEX = re.compile(
    r"V\s*:\s*([+-]?[0-9]*\.?[0-9]+)\s*V\b.*?I\s*:\s*([+-]?[0-9]*\.?[0-9]+)\s*A\b(?:.*?P\s*:\s*([+-]?[0-9]*\.?[0-9]+)\s*W\b)?(?:.*?T\s*:\s*([+-]?[0-9]*\.?[0-9]+)\s*°?C\b)?",
    re.IGNORECASE,
)


def parse_text_samples(text: str) -> List[Dict[str, Any]]:
    """Parse multiline text containing lines like 'V:20.2V I:0.10A P:2.1W' into sample dicts.

    Returns list of dicts with keys: V, I, P (optional), T (optional), t (now).
    Ignores lines that don't match.
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return []

    samples: List[Dict[str, Any]] = []

    # Try CSV with header first (',' or ';')
    delim = ';' if ';' in lines[0] else ','
    header_parts = [p.strip().lower() for p in lines[0].split(delim)]
    if ('v' in header_parts and 'i' in header_parts) or ('voltage' in header_parts and 'current' in header_parts):
        header = header_parts
        for row in lines[1:]:
            parts = [p.strip() for p in row.split(delim)]
            if len(parts) != len(header):
                continue
            data: Dict[str, Any] = {}
            for key, val in zip(header, parts):
                if key in ('t', 'time', 'timestamp'):
                    data['t'] = val
                elif key in ('v', 'voltage'):
                    data['V'] = float(val)
                elif key in ('i', 'current'):
                    data['I'] = float(val)
                elif key in ('p', 'power'):
                    data['P'] = float(val)
                elif key in ('t_c', 'temp', 'temperature'):
                    data['T'] = float(val)
            if 'V' in data and 'I' in data:
                if 'P' not in data:
                    data['P'] = data['V'] * data['I']
                samples.append(data)
        return samples


def parse_csv_bytes(data: bytes) -> List[Dict[str, Any]]:
    """Decode bytes to text (try utf-8/utf-8-sig/latin-1) then reuse parse_text_samples."""
    text = None
    for enc in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        text = data.decode("utf-8", errors="ignore")
    
    # Forcer le traitement comme CSV avec séparateur ; si détecté
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    if not lines:
        return []
    
    # Vérifier si c'est un format CSV avec en-tête et séparateur ;
    if ';' in lines[0]:
        header_parts = [p.strip().lower() for p in lines[0].split(';')]
        samples = []
        for line in lines[1:]:
            if not line.strip():
                continue
            parts = [p.strip() for p in line.split(';')]
            if len(parts) < 2:
                continue
            
            data = {}
            for i, part in enumerate(header_parts):
                if i >= len(parts):
                    break
                val = parts[i]
                if not val:
                    continue
                
                if part in ('v', 'voltage'):
                    try:
                        data['V'] = float(val)
                    except ValueError:
                        continue
                elif part in ('i', 'current'):
                    try:
                        data['I'] = float(val)
                    except ValueError:
                        continue
                elif part in ('p', 'power'):
                    try:
                        data['P'] = float(val)
                    except ValueError:
                        continue
                elif part in ('t_c', 'temp', 'temperature'):
                    try:
                        data['T'] = float(val)
                    except ValueError:
                        continue
            
            if 'V' in data and 'I' in data:
                if 'P' not in data:
                    data['P'] = data['V'] * data['I']
                samples.append(data)
        return samples
    
    # Sinon, utiliser le parser existant
    return parse_text_samples(text)


def parse_xlsx_bytes(data: bytes) -> List[Dict[str, Any]]:
    """Parse an Excel .xlsx file and extract rows as {V,I,P?,T?,t?}.

    Header detection by names: V/Voltage, I/Current, P/Power, T/Temp/Temperature, t/Time/Timestamp.
    If no header row is detected, assume columns A=V, B=I, C=P, D=T.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("openpyxl is required to parse .xlsx files") from e

    wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    def norm(x):
        return str(x).strip().lower() if x is not None else ""

    header_map = {"v": "V", "voltage": "V", "i": "I", "current": "I", "p": "P", "power": "P", "t": "t", "time": "t", "timestamp": "t", "temp": "T", "temperature": "T"}

    # Detect header in first row
    first = rows[0]
    keys = [header_map.get(norm(c), None) for c in first]
    has_header = any(k in ("V", "I", "P", "T", "t") for k in keys)

    samples: List[Dict[str, Any]] = []
    start_idx = 1 if has_header else 0

    if has_header:
        for row in rows[1:]:
            data_row: Dict[str, Any] = {}
            for idx, cell in enumerate(row):
                mapped = keys[idx] if idx < len(keys) else None
                if not mapped:
                    continue
                val = cell
                if mapped in ("V", "I", "P", "T"):
                    try:
                        if val is None or val == "":
                            continue
                        data_row[mapped] = float(val)
                    except Exception:
                        continue
                elif mapped == "t":
                    data_row["t"] = str(val) if val is not None else None
            if "V" in data_row and "I" in data_row:
                if "P" not in data_row:
                    data_row["P"] = data_row["V"] * data_row["I"]
                samples.append(data_row)
    else:
        # Assume A=V, B=I, C=P, D=T
        for row in rows:
            try:
                V = float(row[0]) if len(row) >= 1 and row[0] is not None else None
                I = float(row[1]) if len(row) >= 2 and row[1] is not None else None
                P = float(row[2]) if len(row) >= 3 and row[2] not in (None, "") else None
                T = float(row[3]) if len(row) >= 4 and row[3] not in (None, "") else None
            except Exception:
                continue
            if V is None or I is None:
                continue
            samples.append({"V": V, "I": I, "P": P if P is not None else V * I, "T": T})

    return samples

    # Try plain CSV without header (V,I[,P][,T]) with ',' or ';'
    csv_like = True
    for ln in lines:
        if not re.match(r"^\s*[+-]?[0-9]*\.?[0-9]+\s*[,;]\s*[+-]?[0-9]*\.?[0-9]+(\s*[,;]\s*[+-]?[0-9]*\.?[0-9]+){0,2}\s*$", ln):
            csv_like = False
            break
    if csv_like:
        for ln in lines:
            d = ';' if ';' in ln else ','
            parts = [p.strip() for p in ln.split(d)]
            if len(parts) < 2:
                continue
            V = float(parts[0])
            I = float(parts[1])
            P = float(parts[2]) if len(parts) >= 3 and parts[2] != '' else V * I
            T = float(parts[3]) if len(parts) >= 4 and parts[3] != '' else None
            samples.append({'V': V, 'I': I, 'P': P, 'T': T})
        return samples

    # Fallback to key-value regex lines
    for line in lines:
        m = LINE_REGEX.search(line)
        if not m:
            continue
        V = float(m.group(1))
        I = float(m.group(2))
        P = float(m.group(3)) if m.group(3) is not None else None
        T = float(m.group(4)) if m.group(4) is not None else None
        samples.append({
            'V': V,
            'I': I,
            'P': P if P is not None else V * I,
            'T': T,
            't': datetime.utcnow().isoformat() + 'Z'
        })
    return samples
